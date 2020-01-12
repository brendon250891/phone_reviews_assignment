import created_reviews as cr
import openpyxl
import matplotlib.pyplot as plt
import numpy as np


COMPARISON_EXCEL_WORKBOOK = "comparison.xlsx"
WORKSHEETS = ["reviews per year", "customers"]
REVIEW_HEADINGS = ["Brand", "Product Title", "Year", "Number of Reviews"]
CUSTOMER_HEADINGS = ["Brand", "Percentage of Verified Customers", "Percentage of Each Customer Group"]


def create_new_workbook():
    """ Creates the comparison workbook, sets sheet headings and writes the data to the appropriate sheets. """
    print("Creating Workbook '{0}'".format(COMPARISON_EXCEL_WORKBOOK))
    comparison_workbook = openpyxl.Workbook()
    comparison_workbook.active.title = WORKSHEETS[0]
    comparison_workbook.create_sheet(WORKSHEETS[1])
    insert_worksheet_headings(comparison_workbook[WORKSHEETS[0]], REVIEW_HEADINGS)
    insert_worksheet_headings(comparison_workbook[WORKSHEETS[1]], CUSTOMER_HEADINGS)
    try:
        comparison_workbook.save(COMPARISON_EXCEL_WORKBOOK)
    except PermissionError:
        print("Permission to save the file was denied. Make sure that the file is closed")
    write_data_to_workbook(get_review_yearly_data(), WORKSHEETS[0])
    write_data_to_workbook(get_verified_customer_data(), WORKSHEETS[1])


def insert_worksheet_headings(worksheet, headings):
    """ Inserts the headings into the worksheet.

        Args:
            worksheet (String): The name of the worksheet.
            headings (Array): Array of strings containing the headings to use.
    """
    print("Creating worksheet '{0}'".format(worksheet.title))
    for col in range(1, len(headings) + 1):
        worksheet.cell(row=1, column=col, value="{0}".format(headings[col - 1]))


def write_data_to_workbook(data, sheet_name):
    """ Writes the data retrieved from the database into the correct worksheets

        Args:
            data (Array): Array of tuples retrieved from a database query.
            sheet_name (String): The name of the sheet in which to write the data.
    """
    comparison_workbook = openpyxl.load_workbook(COMPARISON_EXCEL_WORKBOOK)
    if sheet_name not in comparison_workbook.sheetnames:
        comparison_workbook.create_sheet(sheet_name)
    worksheet = comparison_workbook[sheet_name]
    print("Writing '{0}' records into worksheet '{1}'".format(len(data), sheet_name))
    for row in range(2, len(data) + 2):
        for col in range(1, len(data[0]) + 1):
            worksheet.cell(row=row, column=col, value="{0}".format(data[row - 2][col - 1]))
    try:
        comparison_workbook.save(COMPARISON_EXCEL_WORKBOOK)
        comparison_workbook.close()
    except PermissionError:
        print("Permission to save the file was denied. Make sure that the file is closed")


def get_review_yearly_data():
    """ Gets the number of reviews per year for each product, grouping by product title and year

        Returns:
            Array: The records retrieved from the database.
    """
    records = cr.run_database_query("""
        select i.brand, i.title, strftime('%Y', r.review_date) as year, count() from items i 
        join reviews r on r.asin = i.asin group by i.title, year order by i.brand, year""")
    return records


def get_verified_customer_data():
    """ Gets the brands, the percentage of users that have left a review that are verified and the percentage of those in
    relation to all verified users

        Returns:
            Array: The records retrieved from the database.
    """
    records = cr.run_database_query("""
        select i.brand, cast(count() as float) / (select count() from reviews where verified = 'True') * 100,
        cast(count() as float) / (select count() from reviews ir join items ii on ii.asin = ir.asin where ii.brand = i.brand) * 100
        from reviews r join items i on i.asin = r.asin where r.verified = 'True' group by i.brand order by count()
    """)
    return records


def get_workbook_info():
    """ Gets the sheet names and the number of records in each sheet for the workbook if it already exists

        Returns:
            Dictionary: The sheet names and number of records for each sheet.
    """
    workbook_info = {}
    workbook = openpyxl.load_workbook(COMPARISON_EXCEL_WORKBOOK)
    for sheet in workbook.sheetnames:
        workbook_info[sheet] = workbook[sheet].max_row - 1
    return workbook_info


def plot_customer_data(query_data):
    labels = []
    verified_reviews = []
    overall_verified = []
    width = 0.35
    for data in query_data:
        labels.append(data[0])
        overall_verified.append(data[1])
        verified_reviews.append(data[2])
    bar_size = np.arange(len(labels))
    bar1 = plt.bar(bar_size - width / 2, verified_reviews, width, label="Brands Verified User %")
    bar2 = plt.bar(bar_size + width / 2, overall_verified, width, label="Brands Overall Verified User %")
    plt.title("Percentage of Verified Users Grouped By Brand")
    plt.ylabel("Verified Users (%)")
    plt.xlabel("Brand Name")
    plt.yticks(np.arange(0, 110, 10))
    plt.xticks(bar_size, labels=labels)
    plt.legend()
    autolabel(bar1)
    autolabel(bar2)
    plt.show()


def autolabel(bars):
    for bar in bars:
        height = bar.get_height()
        plt.annotate('{:.1f}%'.format(height), xy=(bar.get_x() + bar.get_width() / 2, height), xytext=(2, 1),
                     textcoords="offset points", ha="center", va="bottom")


if __name__ == '__main__':
    """ Executed when the file is run. If the workbook doesn't exist then it is created and data written, otherwise the
    user is prompted with options on how to proceed. """
    if not cr.file_exists(COMPARISON_EXCEL_WORKBOOK):
        create_new_workbook()
    else:
        print()
        print("Workbook '{0}' already exists and contains the following worksheets:".format(COMPARISON_EXCEL_WORKBOOK))
        for sheet, records in get_workbook_info().items():
            print("\t{0} - {1} records".format(sheet, records))
        selection = False
        while not selection:
            user_input = input("Do you want to re-create '{0}' workbook (y/n): ".format(COMPARISON_EXCEL_WORKBOOK)).lower()
            if user_input == 'y':
                create_new_workbook()
                selection = True
            elif user_input == 'n':
                print("No action was taken")
                selection = True
            else:
                print("Invalid selection")
    plot_selection = False
    while not plot_selection:
        user_input = input("Do you want to view customer data graphically? (y/n): ").lower()
        if user_input == 'y':
            plot_customer_data(get_verified_customer_data())
            plot_selection = True
        elif user_input == 'n':
            print("Exiting...")
            plot_selection = True
        else:
            print("Invalid selection")
