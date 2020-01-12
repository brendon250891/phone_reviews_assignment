#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec 19 12:53:54 2019

@author: brendon
"""

import os
import sqlite3
import openpyxl
import re
import datetime

DATABASE_NAME = "phone_reviews_database.db"
ITEMS_EXCEL = "items.xlsx"
REVIEWS_EXCEL = "reviews.xlsx"


def file_exists(filename):
    """ Checks to see if a file already exists.

        Args:
            filename (string): The name of the file to be checked.

        Returns:
            bool: True if the file exists, otherwise false.
    """
    try:
        open(filename)
        return True
    except FileNotFoundError:
        return False


def run_database_query(query, params=None, single_entry=True):
    """ Queries the database

        Args:
            query (string): The query to be run.
            params (tuple, optional): Any parameters the query requires to be executed.
            single_entry (bool, optional): False should be used if inserting multiple rows of data at once. Defaults to
                True.

        Returns:
            Array: The retrieved data of the executed query, emtpy array if query doesn't return values.
    """
    connection = sqlite3.connect(DATABASE_NAME)
    cursor = connection.cursor()
    results = []
    if params is None:
        results = cursor.execute(query).fetchall()
    elif params is not None and single_entry:
        results = cursor.execute(query, params).fetchall()
    elif params is not None and not single_entry:
        cursor.executemany(query, params)
    else:
        results = cursor.execute(query, params).fetchall()
    connection.commit()
    connection.close()
    return results


def create_database():
    """ Creates the database, tables and inserts the data into those tables. """
    print("Creating database...\n")
    run_database_query("drop table if exists items")
    run_database_query("drop table if exists reviews")
    create_items_table()
    create_reviews_table()
    seed_database(ITEMS_EXCEL, 'insert into items values(?,?,?,?,?,?,?,?,?)')
    seed_database(REVIEWS_EXCEL, 'insert into reviews values(NULL,?,?,?,?,?,?,?,?)')


def create_items_table():
    """ Creates the items table for the database """
    run_database_query("""
        create table items(
            asin char(10) not null,
            brand varchar(30) not null,
            title varchar(255) not null,
            url varchar(255) not null,
            image varchar(255) not null,
            rating decimal(1, 1) not null,
            review_url varchar(255) not null,    
            total_reviews integer not null,
            price decimal(4,2) not null,
            constraint items_pk primary key (asin),
            constraint valid_item check (price between '0.00' and '9999.99' and rating between '0.0' and '5.0' and 
                total_reviews >= '0')
        );""")


def create_reviews_table():
    """ Creates the reviews table for the database """
    run_database_query("""
        create table reviews(
            review_id integer,
            asin char(10) not null,
            name varchar(255) not null,
            rating integer(1) not null,
            review_date datetime not null,
            verified boolean default False,
            title varchar(255) not null,
            body text not null, 
            helpful_vote integer not null,
            constraint reviews_pk primary key (review_id)
            constraint reviews_fk foreign key (asin) references items(asin),
            constraint valid_review check (rating between 0 and 5 and date(review_date) is not null and 
                verified in ('True', 'False') and helpful_vote >= 0)
        );""")


def seed_database(excel_file, query):
    """ Seeds the database with the data provided in the excel files.

        Args:
            excel_file (string): The name of the excel file to seed from.
            query (string): The query to be used to insert data.
     """
    data = openpyxl.load_workbook(excel_file)
    print("Workbook '{0}' contains {1} rows of data".format(excel_file, data.active.max_row - 1))
    records_to_seed = prompt_for_number_of_records(data.active.max_row - 1)
    print("\tSeeding data from Workbook '{0}'...".format(excel_file))
    data_to_insert = []
    for row in data.active.iter_rows(min_row=2, max_row=records_to_seed + 1):
        record = []
        for cell in row:
            val = check_for_incorrect_formatting(data.active.cell(row=1, column=cell.column).value, cell.value)
            record.append(val)
        data_to_insert.append(tuple(record))
    run_database_query(query, data_to_insert, False)


def check_for_incorrect_formatting(heading, value):
    """ Fixes some formatting issues in the data sets.
        1. Some data is missing so some defaults are set based on the data type of that column
        2. Some items have multiple prices so the first value is used.
        3. Converts the date column in the reviews dataset to be of type datetime.

        Args:
            heading (string): The column heading
            value (variable type): The value of a cell

        Returns:
             value: The changed value (if it was required), otherwise the value passed in.
    """
    if value is None:
        if re.search('(rating|total|prices|helpful)', heading):
            value = 0
        elif re.search('(title|body)', heading):
            value = 'N/A'
        elif heading == 'name':
            value = 'Anonymous'
    value = str(value)
    if value.startswith('$'):
        value = value[1:value.find(',')]
    elif re.search(", 20[0-9]{2}$", value) and len(value) <= 18:
        year = value[-4:]
        day = value[value.find(' ') + 1: value.find(',')]
        month = month_str_to_int(value[:value.find(' ')].lower())
        value = datetime.datetime(int(year), int(month), int(day))
    return value


def month_str_to_int(month):
    """ Converts a string month to an integer which can handle both long and short months.

        Args:
            month (string): The string representation of the month e.g. Jan or January

        Returns:
            int: An integer representation of the passed in month string
    """
    month_dict = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6, 'jul': 7, 'aug': 8, 'sept': 9, 'oct': 10,
            'nov': 11, 'dec': 12}
    found = re.match("jan|feb|mar|apr|may|jun|jul|aug|sept|nov|dec|oct", month)
    if found is not None:
        return month_dict[month[:found.span()[1]]]
    else:
        return "Invalid Month Given"


def prompt_for_number_of_records(row_count):
    """ Prompts the user for the number of records they wish to insert into the database.

        Args:
            row_count (int): The total number of records in the dataset.

        Returns:
            int: The number of records the user wants to insert.

        Raises:
            ValueError: if the user enters a number outside the range of row_count or an invalid integer a ValueError
                is raised.
     """
    valid_input = False
    while not valid_input:
        user_input = input("\tEnter the number of rows you want to seed (To seed all enter nothing): ")
        if user_input == '':
            valid_input = True
        else:
            try:
                number = int(user_input)
                if 0 <= number <= row_count:
                    row_count = number
                    valid_input = True
                else:
                    raise ValueError
            except ValueError:
                print("Input must be a whole number between 0 and {0}!".format(row_count))
    return row_count


def get_database_info():
    """ Gets the names of the tables that exist in the database.

        Returns:
            Array: Names of the tables within the database.
     """
    return run_database_query("select name from sqlite_master where type = 'table' order by name")


if __name__ == "__main__":
    """ Executed when the file is run. If the database doesn't exist then it is created, otherwise user is prompted with 
    options to select from. """
    if not file_exists(DATABASE_NAME):
        if file_exists(ITEMS_EXCEL) and file_exists(REVIEWS_EXCEL):
            print("No database found with the name '{0}'".format(DATABASE_NAME))
            create_database()
        else:
            print("Database Creation Failed.\nData Files Not Found.")
    else:
        tables = get_database_info()
        print("Database with name '{0}' already exists and contains the following tables:".format(DATABASE_NAME))
        for table in tables:
            number_of_records = run_database_query('select count(*) from {0}'.format(table[0]))
            print("\t{0} -  contains {1} records".format(table[0], number_of_records[0][0]))
        selected_option = False
        while not selected_option:
            user_input = input("\nDo you want to re-create and seed (y/n): ").lower()
            if user_input == 'y':
                os.remove(DATABASE_NAME)
                create_database()
                selected_option = True
            elif user_input == 'n':
                print("Exiting...")
                selected_option = True
            else:
                print("Invalid Selection")
